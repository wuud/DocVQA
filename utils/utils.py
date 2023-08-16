import warnings
import openpyxl
import json
import sys
from fuzzywuzzy import process
sys.path.append('/mnt/workspace/DocVQA/src')
from vectorstore import faiss
from utils import imath

class Config:
    model_path = '/mnt/workspace/DocVQA/resource/指标模版.xlsx'
    json_path = '/mnt/workspace/DocVQA/resource/input.json'
    min_match_score = 80


def loadJsonFile(file_path: str) -> list:
    res = []
    with open(file_path, 'r') as f:
        content = json.load(f)
    for comp in content:
        for year in content[comp]:
            for table in content[comp][year]:
                tb = content[comp][year][table]
                for line in tb:
                    res.append(
                        ','.join([comp, year, table, line, str(tb[line])]))
    return res


def ocrReceived(content :dict) -> list:
    res = []
    for comp in content:
        for year in content[comp]:
                tb = content[comp][year]
                for line in tb:
                    res.append(
                        ','.join([comp, year, line, str(tb[line])]))
    return res


def getScoreList(key: str, lst: list) -> dict:
    res = []
    for raw_key in lst:
        res.append((raw_key, faiss.cos_similarity_search_by_words(
            key, raw_key).tolist()[0][0]))
    res.sort(key=lambda x: x[1], reverse=True)
    return dict(res)


class XlsxReader:
    def __init__(self, fname, shtname):
        with warnings.catch_warnings(record=True):
            self.sheet = openpyxl.load_workbook(fname)[shtname]
            str_ = fname + ' '+shtname + ' read succeed'
            print('\033[1;32m' + str_ + '\033[0m')

    def value(self, x, y):
        if x >= self.sheet.min_row and x <= self.sheet.max_row \
                and y >= self.sheet.min_column and y <= self.sheet.max_column:
            return self.sheet.cell(x, y).value
        else:
            raise Exception(x + ',' + y + ' out of range')


class FinancialIndicators:
    _instance = None
    _flag = False

    def __new__(cls, *args, **kwargs):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self) -> None:
        if not FinancialIndicators._flag:
            FinancialIndicators._flag = True
            file = Config().model_path
            self.balanceSheet = XlsxReader(file, '资产负债表')
            self.incomeSheet = XlsxReader(file, '利润表')
            self.cashFlowSheet = XlsxReader(file, '现金流量表')

    def getValues(self, sheet: XlsxReader, const, begin, end, isCol: bool):
        return [sheet.value(x, const) for x in range(begin, end + 1)] if isCol \
            else [sheet.value(const, y) for y in range(begin, end + 1)]

    # 资产负债表
    def getBalanceIndicators(self) -> list:
        return self.getValues(self.balanceSheet, 1, 2, 63, True)

    # 利润表
    def getIncomeIndicators(self) -> list:
        return self.getValues(self.incomeSheet, 1, 2, 22, True)

    # 现金流量表
    def getCashFlowIndicators(self) -> list:
        return self.getValues(self.cashFlowSheet, 1, 2, 36, True)


class Data:
    _instance = None
    _flag = False

    def __new__(cls, *args, **kwargs):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self, content):
        if not Data._flag:
            Data._flag = True
            #self.data_ = loadJsonFile(Config().json_path)
            self.data_ = ocrReceived(content)

            self.dict_ = dict()
            for line in self.data_:
                v = line
                k = str(line).split(',')[-3]
                if k not in self.dict_.keys():
                    self.dict_[k] = list()
                self.dict_[k].append(v)

            temp = str(self.data_[0]).split(',')

    def getRawDataKeys(self) -> list:
        res = set()
        for line in self.data_:
            res.add(str(line).split(',')[-3])
        return list(res)

    def getSentenceByKey(self, key) -> str:
        if key not in self.dict_.keys():
            raise Exception(str(key)+' is not in dict.keys')
        sentence = str()
        for line in self.dict_[key]:
            s = str(line).split(',')
            sentence = sentence + \
                '{}的{}是{}元。'.format(s[-2], s[-3], s[-1])
        return sentence

    def getRawListByKey(self, key) -> list:
        if key not in self.dict_.keys():
            raise Exception(str(key)+' is not in dict.keys')
        return self.dict_[key]


class MatchUint:
    def __init__(self, rawKey, number1, number2):
        self.rawKey = rawKey
        self.num1 = number1
        self.num2 = number2

    def __str__(self):
        return 'key:' + self.rawKey + ',' \
            + 'num1:' + str(self.num1) + ',' \
            + 'num2:' + str(self.num2)


class Context:
    _instance = None
    _flag = False

    def __new__(cls, *args, **kwargs):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self, content):
        if not Context._flag:
            Context._flag = True
            self.data_ = Data(content)
            self.indicators_ = FinancialIndicators()

            raw_keys = self.data_.getRawDataKeys()
            models = self.indicators_.getBalanceIndicators(
            ) + self.indicators_.getIncomeIndicators()

            self.modelsMap_ = {}

            for model in models:
                match_result = process.extract(model, raw_keys)[0]
                print(model,'--------匹配---------',match_result[0],match_result[1])
                if match_result[1] >= Config().min_match_score:
                    k = match_result[0]
                    lst = self.data_.getRawListByKey(k)
                    nums = []
                    for line in lst:
                        nums.append(float(str(line).split(',')[-1]))
                    print('  len(nums):',len(nums))
                    if len(nums) == 2:
                        self.modelsMap_[model] = MatchUint(k, nums[0], nums[1])

            #data、context添加二级指标
            model_keys_ = self.modelsMap_.keys()
            if '资产总计' in model_keys_ and '净利润' in model_keys_:
                rate = self.get('净利润')
                zichan = self.get('资产总计')
                zcJLV1,zcJLV2 = imath.zc_net_interest_rates(rate.num1, zichan.num1 ),imath.zc_net_interest_rates(rate.num2, zichan.num2)
                self.modelsMap_['资产净利率'] = MatchUint('资产净利率', zcJLV1, zcJLV2)

                self.data_.data_.append('资产净利率,本期金额,' + str(zcJLV1))
                self.data_.data_.append('资产净利率,上期金额,' + str(zcJLV2))

                self.data_.dict_['资产净利率'] = list()
                self.data_.dict_['资产净利率'].append('资产净利率,本期金额,' + str(zcJLV1))
                self.data_.dict_['资产净利率'].append('资产净利率,上期金额,' + str(zcJLV2))


            if '流动资产合计' in model_keys_ and '流动负债合计' in model_keys_:
                zc = self.get('流动资产合计')
                fz = self.get('流动负债合计')
                n1,n2 = imath.qy_net_interest_rates(zc.num1, fz.num1),imath.qy_net_interest_rates(zc.num2, fz.num2)
                self.modelsMap_['流动比率'] = MatchUint('流动比率', n1, n2)

                self.data_.data_.append('流动比率,本期金额,' + str(n1))
                self.data_.data_.append('流动比率,上期金额,' + str(n2))

                self.data_.dict_['流动比率'] = list()  
                self.data_.dict_['流动比率'].append('流动比率,本期金额,' + str(n1))
                self.data_.dict_['流动比率'].append('流动比率,上期金额,' + str(n2))


    def get(self, key) -> MatchUint:
        return self.modelsMap_[key]
